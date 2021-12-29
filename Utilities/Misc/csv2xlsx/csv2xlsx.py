"""csv2xlsx cut and paste

This script opens a series of comma-delimited .csv files and copies portions of columns into specified
worksheets of a specified Excel (.xlsx) file. It saves the Excel file with a new name. The names of the
new and old Excel files are currently hardwired in the script below.

The names of the csv files are listed in the file called 'csv2xlxs_parameters.csv'. This file contains the full
path to the csv files, followed by the row/column pair where the data starts, the length of the column
snippet, the name of the worksheet in the Excel file, and the row/column pair where the snippet is to be
pasted.

Note that the module used to save the new Excel file does not support wmf (Windows MetaFile) images. Any
wmf images in the original Excel file will be dropped from the new Excel file.

This script requires the Excel parsing module 'openpyxl' and the csv reader/writer 'csv'.
"""

import csv
from openpyxl import load_workbook

wb = load_workbook('MEPFires_database_comparison_draft.xlsx')

param_file = open('csv2xlsx_parameters.csv','r')
param_reader = csv.reader(param_file,delimiter=',')
params = list(param_reader)

for i in range(0,len(params)):

    orig_row = int(params[i][1]) - 1
    orig_col = int(params[i][2]) - 1
    csv_file = open(params[i][0],'r')
    data = list(csv.reader(csv_file,delimiter=','))
    new_row = int(params[i][5])
    new_col = int(params[i][6])
    length  = int(params[i][3])
    print(i,new_row,new_col)
 
    ws = wb[params[i][4]]

    for counter in range(0,abs(length)):
        if length > 0:
            _ = ws.cell(row=new_row+counter, column=new_col, value=round(float(data[orig_row+counter][orig_col]),2))
        else:
            print(counter)
            _ = ws.cell(row=new_row+counter, column=new_col, value=round(float(data[orig_row][orig_col+counter]),2))

    csv_file.close()

wb.save('MEPFires_database_comparison_draft_modified.xlsx')
